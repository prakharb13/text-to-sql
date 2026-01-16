from pathlib import Path
from openpyxl import Workbook


def save_results_to_excel(prompt_results, model_name, original_results=None, custom_results=None):
    """
    Save evaluation results to Excel file with summary statistics.
    
    Args:
        prompt_results: Dictionary mapping keys (model_key_prompt_name) to lists of results
                       Each result dict has 'sql_match' and 'answer_match'
        model_name: Name identifier for the file
        original_results: Optional dictionary for original test case results
        custom_results: Optional dictionary for custom test case results
    """
    # Create output folder
    output_folder = Path("model_output")
    output_folder.mkdir(exist_ok=True)
    
    # Create Excel workbook
    wb = Workbook()
    
    # Summary sheet (all test cases combined)
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Model Name", "Prompt Type", "Avg SQL Match Score", "Avg Answer Match Score"])
    
    for key, results in prompt_results.items():
        if results:
            model_key, prompt_name = key.split("__", 1)
            avg_sql_match = sum(r['sql_match'] for r in results) / len(results)
            avg_answer_match = sum(r['answer_match'] for r in results) / len(results)
            summary_sheet.append([model_key, prompt_name, avg_sql_match, avg_answer_match])
    
    # Original test cases sheet
    if original_results:
        original_sheet = wb.create_sheet("Original Test Cases")
        original_sheet.append(["Model Name", "Prompt Type", "Avg SQL Match Score", "Avg Answer Match Score"])
        
        for key, results in original_results.items():
            if results:
                model_key, prompt_name = key.split("__", 1)
                avg_sql_match = sum(r['sql_match'] for r in results) / len(results)
                avg_answer_match = sum(r['answer_match'] for r in results) / len(results)
                original_sheet.append([model_key, prompt_name, avg_sql_match, avg_answer_match])
    
    # Custom test cases sheet
    if custom_results:
        custom_sheet = wb.create_sheet("Custom Test Cases")
        custom_sheet.append(["Model Name", "Prompt Type", "Avg SQL Match Score", "Avg Answer Match Score"])
        
        for key, results in custom_results.items():
            if results:
                model_key, prompt_name = key.split("__", 1)
                avg_sql_match = sum(r['sql_match'] for r in results) / len(results)
                avg_answer_match = sum(r['answer_match'] for r in results) / len(results)
                custom_sheet.append([model_key, prompt_name, avg_sql_match, avg_answer_match])
    
    # Save file
    output_file = output_folder / f"text_to_sql_model_results.xlsx"
    wb.save(output_file)
    print(f"\nResults saved to: {output_file}")
    if original_results:
        print(f"  - Summary sheet: All test cases")
        print(f"  - Original Test Cases sheet: {len([r for r in original_results.values() if r])} model-prompt combinations")
    if custom_results:
        print(f"  - Custom Test Cases sheet: {len([r for r in custom_results.values() if r])} model-prompt combinations")
